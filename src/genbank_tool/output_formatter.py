"""Output formatting for sequence data."""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from genbank_tool.models import RetrievedSequence
from genbank_tool.transcript_selector import TranscriptSelection
from genbank_tool.data_validator import ValidationResult

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class OutputFormatter:
    """Formatter for structured output with Excel compatibility."""
    
    # Column headers
    COLUMNS = [
        "Input Name",
        "Official Symbol", 
        "Gene ID",
        "RefSeq Accession",
        "GenBank URL",
        "CDS Length",
        "CDS Sequence",
        "Selection Method",
        "Confidence Score",
        "Warnings",
        "Validation Status",
        "Validation Confidence",
        "Validation Issues",
        "Error"
    ]
    
    def __init__(self, include_audit_trail: bool = True):
        """
        Initialize the formatter.
        
        Args:
            include_audit_trail: Whether to generate audit trail
        """
        self.include_audit_trail = include_audit_trail
        self.audit_data = {
            'start_time': datetime.now(),
            'entries': [],
            'statistics': {},
            'database_versions': {}
        }
    
    def format_results(self, 
                      results: List[Dict[str, Any]],
                      output_path: Union[str, Path],
                      format: str = 'tsv',
                      excel_compatible: bool = True) -> None:
        """
        Format and write results to file.
        
        Args:
            results: List of result dictionaries
            output_path: Path to output file
            format: Output format ('tsv', 'csv', 'json', 'excel')
            excel_compatible: Use UTF-8 BOM for Excel compatibility
        """
        path = Path(output_path)
        
        if format == 'tsv':
            self._write_tsv(results, path, excel_compatible)
        elif format == 'csv':
            self._write_csv(results, path, excel_compatible)
        elif format == 'json':
            self._write_json(results, path)
        elif format == 'excel':
            self._write_excel(results, path)
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        # Write audit trail if enabled
        if self.include_audit_trail:
            self._write_audit_trail(path)
    
    def format_sequence_result(self,
                             input_name: str,
                             sequence: Optional[RetrievedSequence] = None,
                             selection: Optional[TranscriptSelection] = None,
                             validation: Optional[ValidationResult] = None,
                             error: Optional[str] = None) -> Dict[str, Any]:
        """
        Format a single sequence result.
        
        Args:
            input_name: Original input gene name
            sequence: Retrieved sequence data
            selection: Transcript selection info
            validation: Validation results
            error: Error message if failed
            
        Returns:
            Formatted result dictionary
        """
        result = {
            'Input Name': input_name,
            'Official Symbol': '',
            'Gene ID': '',
            'RefSeq Accession': '',
            'GenBank URL': '',
            'CDS Length': '',
            'CDS Sequence': '',
            'Selection Method': '',
            'Confidence Score': '',
            'Warnings': '',
            'Validation Status': '',
            'Validation Confidence': '',
            'Validation Issues': '',
            'Error': error or ''
        }
        
        if sequence:
            result.update({
                'Official Symbol': sequence.gene_symbol,
                'Gene ID': sequence.gene_id,
                'RefSeq Accession': f"{sequence.accession}.{sequence.version}",
                'GenBank URL': sequence.genbank_url,
                'CDS Length': str(sequence.cds_length),
                'CDS Sequence': sequence.cds_sequence
            })
        
        if selection:
            # Handle SelectionMethod enum
            method = selection.method.value if hasattr(selection.method, 'value') else str(selection.method)
            result.update({
                'Selection Method': method,
                'Confidence Score': f"{selection.confidence:.2f}",
                'Warnings': '; '.join(selection.warnings) if selection.warnings else ''
            })
        
        if validation:
            # Format validation status
            if validation.is_valid:
                status = 'Valid'
            else:
                status = 'Invalid'
            
            # Format validation issues
            issues = []
            for issue in validation.issues:
                issue_str = f"[{issue.level.value}] {issue.message}"
                issues.append(issue_str)
            
            result.update({
                'Validation Status': status,
                'Validation Confidence': f"{validation.confidence_score:.2f}",
                'Validation Issues': '; '.join(issues) if issues else ''
            })
        
        # Update audit data
        if self.include_audit_trail:
            # Handle SelectionMethod enum for JSON serialization
            selection_method = None
            if selection:
                selection_method = selection.method.value if hasattr(selection.method, 'value') else str(selection.method)
            
            self.audit_data['entries'].append({
                'input_name': input_name,
                'timestamp': datetime.now().isoformat(),
                'success': error is None,
                'selection_method': selection_method,
                'validation_status': validation.is_valid if validation else None
            })
        
        return result
    
    def _write_tsv(self, results: List[Dict[str, Any]], path: Path, excel_compatible: bool) -> None:
        """Write TSV file with optional UTF-8 BOM."""
        encoding = 'utf-8-sig' if excel_compatible else 'utf-8'
        
        with open(path, 'w', encoding=encoding, newline='') as f:
            writer = csv.DictWriter(f, fieldnames=self.COLUMNS, delimiter='\t')
            writer.writeheader()
            writer.writerows(results)
    
    def _write_csv(self, results: List[Dict[str, Any]], path: Path, excel_compatible: bool) -> None:
        """Write CSV file with optional UTF-8 BOM."""
        encoding = 'utf-8-sig' if excel_compatible else 'utf-8'
        
        with open(path, 'w', encoding=encoding, newline='') as f:
            writer = csv.DictWriter(f, fieldnames=self.COLUMNS)
            writer.writeheader()
            writer.writerows(results)
    
    def _write_json(self, results: List[Dict[str, Any]], path: Path) -> None:
        """Write JSON file."""
        output = {
            'metadata': {
                'generated': datetime.now().isoformat(),
                'total_entries': len(results),
                'columns': self.COLUMNS
            },
            'results': results
        }
        
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
    
    def _write_excel(self, results: List[Dict[str, Any]], path: Path) -> None:
        """Write Excel file with formatting."""
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel output requires openpyxl. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sequence Results"
        
        # Write headers with formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col, header in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for row, result in enumerate(results, 2):
            for col, header in enumerate(self.COLUMNS, 1):
                value = result.get(header, '')
                ws.cell(row=row, column=col, value=value)
                
                # Color code based on status
                if header == 'Validation Status':
                    cell = ws.cell(row=row, column=col)
                    if value == 'Valid':
                        cell.font = Font(color="008000")  # Green
                    elif value == 'Invalid':
                        cell.font = Font(color="FF0000")  # Red
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Add metadata sheet
        if self.include_audit_trail:
            meta_ws = wb.create_sheet("Metadata")
            meta_ws.append(["Generated", datetime.now().isoformat()])
            meta_ws.append(["Total Entries", len(results)])
            meta_ws.append(["Successful", sum(1 for r in results if not r.get('Error'))])
            meta_ws.append(["Failed", sum(1 for r in results if r.get('Error'))])
        
        wb.save(path)
    
    def _write_audit_trail(self, output_path: Path) -> None:
        """Write audit trail file."""
        audit_path = output_path.with_suffix('.audit.json')
        
        # Calculate statistics
        total = len(self.audit_data['entries'])
        successful = sum(1 for e in self.audit_data['entries'] if e['success'])
        
        self.audit_data.update({
            'end_time': datetime.now(),
            'statistics': {
                'total_processed': total,
                'successful': successful,
                'failed': total - successful,
                'success_rate': f"{(successful/total*100):.1f}%" if total > 0 else "0%"
            },
            'database_versions': {
                'ncbi_gene': 'Current',
                'ncbi_refseq': 'Current',
                'uniprot': 'Current',
                'ensembl': 'Current'
            }
        })
        
        # Selection method breakdown
        method_counts = {}
        for entry in self.audit_data['entries']:
            method = entry.get('selection_method', 'None')
            method_counts[method] = method_counts.get(method, 0) + 1
        
        self.audit_data['statistics']['selection_methods'] = method_counts
        
        # Validation status breakdown
        validation_counts = {'valid': 0, 'invalid': 0, 'not_validated': 0}
        for entry in self.audit_data['entries']:
            status = entry.get('validation_status')
            if status is True:
                validation_counts['valid'] += 1
            elif status is False:
                validation_counts['invalid'] += 1
            else:
                validation_counts['not_validated'] += 1
        
        self.audit_data['statistics']['validation_results'] = validation_counts
        
        # Write audit file
        with open(audit_path, 'w', encoding='utf-8') as f:
            json.dump(self.audit_data, f, indent=2, default=str)
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get current processing statistics."""
        total = len(self.audit_data['entries'])
        successful = sum(1 for e in self.audit_data['entries'] if e['success'])
        
        return {
            'total_processed': total,
            'successful': successful,
            'failed': total - successful,
            'duration': str(datetime.now() - self.audit_data['start_time'])
        }